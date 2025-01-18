#!!!!Important note: before convert firstly only keep first page in excel, remove table header and left pages 

import pandas as pd
from bs4 import BeautifulSoup
import argparse
import logging
import os
from openpyxl import load_workbook

logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

def parse_args():
    parser = argparse.ArgumentParser(description="Replace HTML table cell values with data from an Excel file.")
    parser.add_argument('excel_file', type=str, help='Path to the Excel file')
    return parser.parse_args()

def read_excel_with_merged_cells(excel_file):
    wb = load_workbook(excel_file, data_only=True)
    ws = wb.active
    data = []

    for row in ws.iter_rows(values_only=True):
        data.append([cell for cell in row])

    # Handle merged cells
    for merge in ws.merged_cells.ranges:
        min_col = merge.min_col - 1
        min_row = merge.min_row - 1
        max_col = merge.max_col - 1
        max_row = merge.max_row - 1
        merge_value = data[min_row][min_col]
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                if row == min_row and col == min_col:
                    continue
                data[row][col] = None

    logging.debug(f'----------------------{data}-----------------')
    # Remove columns that are entirely None
    df = pd.DataFrame(data)
    df.dropna(axis=1, how='all', inplace=True)
    logging.debug(f'----------------------{df}------------------')

    return df

def main():
    # Parse command-line arguments
    args = parse_args()

    base_name = os.path.splitext(args.excel_file)[0]
    output_file = f"{base_name}.html"

    # Read the Excel file with merged cells
    excel_df = read_excel_with_merged_cells(args.excel_file)

    logging.debug(f'Excel DataFrame columns:\n, {excel_df.columns}')

    # Read the existing HTML file
    with open('template.html', 'r', encoding='utf-8') as file:
        soup = BeautifulSoup(file, 'html.parser')

    # Find the HTML table
    table = soup.find('table')

    if table:
        for row_num in range(11, 52):  # HTML rows 11 to 51
            tr = table.find('tr', {'rn': str(row_num)})
            if tr:
                # Get the corresponding data row from the Excel file
                data_row = excel_df.iloc[row_num - 11]
                # Find all <td> elements in the <tr>
                tds = tr.find_all('td')
                # Replace the values in the <td> elements with the values from the Excel row
                # Ensure the indices exist in data_row before accessing them
                indices_to_replace = excel_df.columns[-4:]
                existing_indices = [i for i in indices_to_replace if i in data_row.index]
                logging.debug(f'existing_indices: {existing_indices}')

                # Extract values from the existing indices
                values_to_replace = data_row[existing_indices].values

                # Target the last 5 <td> elements
                for i, td in enumerate(tds[-4:]):
                    # Format the value to include commas in numbers
                    value = values_to_replace[i]
                    logging.debug(f'value: {value}')
                    if isinstance(value, (int, float)):
                        td.string = f'{value:,.2f}'
                    else:
                        td.string = str(value)
            else:
                logging.error(f"No <tr> element with rn='{row_num}' found")
                
    else:
        logging.error("No HTML table found")

    # Save the modified HTML file with formatted output
    with open(output_file, 'w', encoding='utf-8') as file:
        # Write each <td> element on a new line
        formatted_html = str(soup).replace('</td><td>', '</td>\n<td>').replace('</tr><tr>', '</tr>\n<tr>')
        file.write(formatted_html)

    logging.info("HTML file has been updated")

if __name__ == "__main__":
    main()
